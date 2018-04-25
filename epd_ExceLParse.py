# -*- coding: UTF-8 -*-
import datetime
# import sys
import openpyxl
import os
import shutil
from epd_util import Epd_util
from configUtil import ConfigUtil
import pyodbc

# epd_batchinfo
# epd_batchstatement
# epd_pginfo
# epd_expertinfo
# epd_projectinfo
# epd_log


def loadExcel():
    '''
    读取配置文件中待解析excel文件\文件列表,以及输出sql文件路径.
    解析Excel内容,生成Sql Insert语句执行,并保存到sql文件留档
    '''
    config = ConfigUtil()
    filenamepath, outputpath = config.getConfigString(
        'EXCELINPUTPATH'), config.getConfigString('SQLOUTPUTPATH')

    if type(filenamepath) is list:
        for filename in filenamepath:
            print('解析配置项: %s' % (filename))
            loadExcelData(filename, outputpath)
            print('解析完毕', end='\n\n')
            pass
        pass
    elif type(filenamepath) is str:
        loadExcelData(filenamepath, outputpath)
        pass


def loadExcelData(filenamepath, outputpath):
    (filename, ext) = os.path.splitext(os.path.basename(filenamepath))
    wb = openpyxl.load_workbook(filenamepath)
    foldname = '%s_%s生成' % (filename,
                            str(datetime.datetime.now().strftime('%Y%m%d')))
    sqlfilenamepath = outputpath + foldname + '/'
    if os.path.exists(sqlfilenamepath):
        shutil.rmtree(sqlfilenamepath)
    os.makedirs(sqlfilenamepath)
    batchid = wb['epd_batchinfo']['A3'].value
    batchgroupid = wb['epd_batchinfo']['F3'].value
    for sheet in wb.sheetnames:
        print('解析表%s开始 batchgroupid: %s' % (sheet, batchgroupid))
        parseSheetData(wb[sheet], sqlfilenamepath, batchid, batchgroupid)

    pass


def parseSheetData(sheet, sqlfilenamepath, batchid, batchgroupid):
    epd_u, config = Epd_util(), ConfigUtil()
    sheet_title = sheet.title
    columnpart, valuepart = epd_u.BASE_SQL[sheet_title], epd_u.VALUE_SQL[
        sheet_title]
    maxRow, maxCol = sheet.max_row, sheet.max_column
    outputfilename = sheet_title + '_' + str(
        datetime.datetime.now().strftime('%Y%m%d%H%M')) + '.sql'
    cnxn = pyodbc.connect(config.getConfigString('CONNECTSTRING', 'DATABASE'))
    cursor = cnxn.cursor()
    with open(sqlfilenamepath + outputfilename, 'w', encoding='utf-8') as fw:
        for row in sheet.iter_rows(row_offset=2, max_row=maxRow - 2):
            rowlist = []
            for cell in row:
                cellvalue = cell.value
                # if not cellvalue: is None 判断是否None
                if cellvalue is None:
                    rowlist.append('')
                else:
                    rowlist.append(cellvalue)
            if sheet_title == 'epd_batchstatement':
                rowlist.append(batchgroupid)
            if sheet_title == 'epd_pginfo':
                rowlist.append(batchid)
                rowlist.append(batchgroupid)

            mainsql = columnpart + valuepart.format(rowlist)
            fw.write(mainsql + '\n')
            cursor.execute(mainsql)
        fw.write('--%s数据共计%d条\n\n' % (sheet_title, maxRow - 2))
    cnxn.commit()
    pass


if __name__ == "__main__":
    # loadExcelData(sys.argv[1])
    # loadExcelData(
    #     'D:/cvsdocument/应用开发部\科研计划项目\专家评审费发放管理系统/维护文档/2018年度/万达认定数据导入/原始备份/高企财务导数据2016.xlsx',
    #     'D:/cvsdocument/应用开发部/科研计划项目/专家评审费发放管理系统/维护文档/2018年度/万达认定数据导入/导入脚本/')
    loadExcel()
    pass
